import numpy as np
import datetime
import calendar
import pickle
import base64
import pandas as pd
import streamlit as st
import time
import threading
from typing import Literal
from openpyxl import load_workbook
import streamlit.components.v1 as components
from streamlit_shortcuts import add_keyboard_shortcuts
from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx

notification_container = None
DATA_VERSION = 3

# Data Schema
projects_schema = {
    "name": "string",
    "start_date": "datetime64[ns]",
    "end_date": "datetime64[ns]",
    "executed": "datetime64[ns]"
}

persons_schema = {
    "name": "string",
    "gender": "string"
}

activities_schema = {
    "project": "string",
    "wp": "string",
    "activity": "string",
    "trl": "string",
    "start_date": "datetime64[ns]",
    "end_date": "datetime64[ns]",
    "real_start_date": "datetime64[ns]",
    "real_end_date": "datetime64[ns]"
}

contracts_schema = {
    "project": "string",
    "person": "string",
    "profile": "string",
    "gender": "string",
    "start_date": "datetime64[ns]",
    "end_date": "datetime64[ns]",
}

working_days_schema = {
    "project": "string",
    "date": "datetime64[ns]",
    "day": "int64"
}

sheets_schema = {
    "person": "string",
    "date": "datetime64[ns]",
    "Jornada Di√°ria": "int64",
    "Faltas": "float64",
    "F√©rias": "float64",
    "Sal√°rio": "float64",
    "SS": "float64",
}

real_work_schema = {
    "person": "string",
    "project": "string",
    "date": "datetime64[ns]",
    "hours": "int64"
}

planned_work_schema = {
    "person": "string",
    "project": "string",
    "activity": "string",
    "date": "datetime64[ns]",
    "hours": "int64"
}

inv_order_num_schema = {
    'project': 'string',
    "wp": "string",
    "trl": "string",
    "code": "string"
}


def create_session(reset=False):
    """ Initialize streamlit session variables """

    if 'key' not in st.session_state:
        st.session_state.key = 0

    if 'to_reset' not in st.session_state:
        st.session_state.to_reset = False

    if 'notification' not in st.session_state:
        st.session_state.notification = None

    if 'unsaved' not in st.session_state or reset:
        st.session_state.unsaved = False

    if 'file_id' not in st.session_state or reset:
        if reset:
            st.session_state.file_id = (st.session_state.file_id + 1) % 2
        else:
            st.session_state.file_id = 0

    if 'company_name' not in st.session_state or reset:
        st.session_state.company_name = ""

    if 'activities' not in st.session_state or reset:
        st.session_state.activities = pd.DataFrame(
            columns=activities_schema.keys()).astype(activities_schema)
        st.session_state.persons = pd.DataFrame(
            columns=persons_schema.keys()).astype(persons_schema)
        st.session_state.contracts = pd.DataFrame(
            columns=contracts_schema.keys()).astype(contracts_schema)
        st.session_state.projects = pd.DataFrame(
            columns=projects_schema.keys()).astype(projects_schema)
        st.session_state.sheets = pd.DataFrame(
            columns=sheets_schema.keys()).astype(sheets_schema)
        st.session_state.planned_work = pd.DataFrame(
            columns=planned_work_schema.keys()).astype(planned_work_schema)
        st.session_state.real_work = pd.DataFrame(
            columns=real_work_schema.keys()).astype(real_work_schema)
        st.session_state.working_days = pd.DataFrame(
            columns=working_days_schema.keys()).astype(working_days_schema)
        st.session_state.inv_order_num = pd.DataFrame(
            columns=inv_order_num_schema.keys()).astype(inv_order_num_schema)


def save_data():
    """ Download all data """

    object_to_download = pickle.dumps({
        'activities': st.session_state.activities,
        'contracts': st.session_state.contracts,
        'projects': st.session_state.projects,
        'sheets': st.session_state.sheets,
        'planned_work': st.session_state.planned_work,
        'real_work': st.session_state.real_work,
        'working_days': st.session_state.working_days,
        'persons': st.session_state.persons,
        'inv_order_num': st.session_state.inv_order_num,
        'version': DATA_VERSION
    })

    b64 = base64.b64encode(object_to_download).decode()
    download_filename = st.session_state.company_name+".pkl"

    components.html(
        f"""
            <html>
                <head>
                <title>Start Auto Download file</title>
                <a id="fileDownload" href="data:application/octet-stream;base64,{b64}" download="{download_filename}">
                <script>
                    document.getElementById('fileDownload').click();
                </script>
                </head>
            </html>
        """,
        height=0,
    )

    st.session_state.unsaved = False


def save_excel(file, name):
    b64 = base64.b64encode(file).decode()

    components.html(
        f"""
            <html>
                <head>
                <title>Start Auto Download file</title>
                <a id="fileDownload" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{name}">
                <script>
                    document.getElementById('fileDownload').click()
                </script>
                </head>
            </html>
        """,
        height=0,
    )


def invalid(*args):
    """ Check if variables aren't empty """

    for arg in args:
        if arg is None:
            return True

        if type(arg) == str and len(arg) == 0:
            return True

    return False


def date_range(start, end):
    months_range = pd.date_range(start=get_first_date(
        start), end=get_last_date(end), freq='MS')
    return [month.strftime('%b/%y') for month in months_range]


def reset_key():
    st.session_state.key = (st.session_state.key + 1) % 2
    st.rerun()


def get_first_date(date):
    if not date:
        return None

    return datetime.date(int(date.year), int(date.month), 1)


def get_last_date(date):
    _, last_day = calendar.monthrange(date.year, date.month)
    return datetime.date(date.year, date.month, last_day)


def extract_cell_colors_and_dates(file):
    # Load the Excel workbook
    wb = load_workbook(file, data_only=True)
    ws = wb['Cronograma']

    months = pd.read_excel(file, sheet_name="Cronograma",
                           header=8, nrows=0).iloc[:, 6:]

    # Initialize an empty list to store cell colors
    colors_data = []

    # Iterate through each row and column in the worksheet

    for row in ws.iter_rows():
        row_colors = []
        for cell in row:
            # Get the fill color of the cell
            fill = cell.fill.start_color.index
            row_colors.append(fill)
        colors_data.append(row_colors)

    # Convert the list of colors into a DataFrame
    df = pd.DataFrame(colors_data)
    df = df.iloc[:, 6: 6+len(months.columns)]

    active_color = ws['G1325'].fill.start_color.index
    deactivated_color = ws['O1326'].fill.start_color.index
    extended_color = ws['G1326'].fill.start_color.index

    df = df.replace(active_color, 1)
    df = df.replace(deactivated_color, -1)
    df = df.replace(extended_color, 2)

    df.columns = months.columns

    return df, months.columns[0], months.columns[-1]


def min_max_dates(row, value1, value2):
    dates = []
    for col, val in row.items():
        if val in [value1, value2]:
            dates.append(col)
    if len(dates) == 0:
        return None, None
    return get_first_date(min(dates)), get_last_date(max(dates))


@st.experimental_dialog("Pretende Continuar?", width="large")
def get_dialog(title: str, paragraph: str, action):
    st.title(title)
    st.write(paragraph)

    c1, c2 = st.columns(2)
    if c1.button("Continuar", use_container_width=True):
        action()
        st.rerun()

    if c2.button("Cancelar", use_container_width=True):
        st.rerun()


def get_topbar(title: str, buttons=True) -> tuple[bool, bool] | None:
    with st.container(border=True):

        c1, c2 = st.columns([0.7, 0.3] if buttons else [0.9999, 0.0001])

        c1.header(title)

        if buttons:
            with c2:
                if save := st.button(":floppy_disk: Guardar", use_container_width=True):
                    st.session_state.unsaved = True
                undo = st.button(":x: Desfazer", use_container_width=True)

            return save, undo


def set_notification(type: Literal['warning', 'error', 'success'], message: str, force_reset=False):

    st.session_state.notification = {"message": message, "type": type}

    if force_reset:
        return reset_key()

    global notification_container

    match type:
        case 'warning':
            notification_container.warning(message, icon="‚ö†Ô∏è")
        case 'error':
            notification_container.error(message, icon="üö®")
        case'success':
            notification_container.success(message, icon="‚úÖ")


def check_notification():
    global notification_container

    st.markdown("""
        <style>
            div.stAlert > div > div > div > div{
                display: flex;
                flex-direction: row;
                justify-content: center;
            }
            
            div.stAlert {
                position: fixed;
                top: 70px;
                left: 50%;
                z-index: 50;
            }
        </style>
    """, unsafe_allow_html=True)

    notification_container = st.empty()

    if notification := st.session_state.notification:

        match notification['type']:
            case 'warning':
                notification_container.warning(
                    notification['message'], icon="‚ö†Ô∏è")
            case 'error':
                notification_container.error(notification['message'], icon="üö®")
            case'success':
                notification_container.success(
                    notification['message'], icon="‚úÖ")


def fade_notification():
    global notification_container

    if st.session_state.notification:
        time.sleep(2)
        notification_container.empty()

        notification_container = None
        st.session_state.notification = None


def warning_before_leave():
    if st.session_state.unsaved:
        js = """
        <script>
        window.onbeforeunload = function() {
            return "Do you really want to leave this page and lose your unsaved changes?";
        };
        </script>
        """
        components.html(js, height=0)


def sync_dataframes():
    if 'run' not in st.session_state:
        st.session_state.run = 0
    else:
        st.session_state.run = (st.session_state.run + 1) % 2

    js = f"""
        <script>
        hash = "{st.session_state.run}";
        tables = window.parent.document.querySelectorAll('.dvn-scroller');

        isSyncingScroll = false;

        tables.forEach((table, index) => {{
            table.addEventListener('scroll', function() {{
                if (!isSyncingScroll) {{
                    isSyncingScroll = true;
                    tables.forEach((otherTable, otherIndex) => {{
                        if (index !== otherIndex) {{
                            otherTable.scrollLeft = table.scrollLeft;
                        }}
                    }});
                    isSyncingScroll = false;
                }}
            }});
        }});
        </script>
    """

    components.html(js, height=0)


def compare(date_to, lower_date=None, higher_date=None):

    if not lower_date and not higher_date:
        return None

    def to_first_of_month(d):
        if not d:
            return None

        if isinstance(d, pd.Timestamp):
            return datetime.date(d.year, d.month, 1)
        elif isinstance(d, datetime.datetime):
            return datetime.date(d.year, d.month, 1)
        elif isinstance(d, datetime.date):
            return datetime.date(d.year, d.month, 1)
        else:
            raise TypeError("Unsupported date type")

    date_to = to_first_of_month(date_to)
    lower_date = to_first_of_month(lower_date)
    higher_date = to_first_of_month(higher_date)

    if lower_date and date_to < lower_date:
        return False

    if higher_date and date_to > higher_date:
        return False

    return True


def floor_map(x, decimal_places):
    if isinstance(x, float):
        factor = 10 ** decimal_places
        return np.floor(x * factor) / factor
    return x


def round_down(serie, decimal_places=0):
    fator = 10 ** decimal_places

    # Apply flooring only to non-NaN numbers that are floats
    def floor_if_float(x):
        if pd.isna(x):
            return np.nan
        elif isinstance(x, float):
            return np.floor(x * fator) / fator
        else:
            return x  # Leave integers unchanged

    return serie.apply(floor_if_float)


def selected_operation():
    html_code = """
    
    <div id="statistics">
        <div class="stats-header">Statistics</div>
        <div class="stats-row">
            <div class="stats-cell">Max:</div>
            <div class="stats-cell" id="maxValue">0.00</div>
            <div class="stats-cell">Min:</div>
            <div class="stats-cell" id="minValue">0.00</div>
            <div class="stats-cell">Sum:</div>
            <div class="stats-cell" id="sumValue">0.00</div>
            <div class="stats-cell">Mean:</div>
            <div class="stats-cell" id="meanValue">0.00</div>
        </div>
    </div>

    <script>
        let calculationTimeout;
        const doc = window.parent.document;

        const xpath = "(//div[@data-testid='element-container'])[last()]";
        const container = doc.evaluate(xpath, doc, null, XPathResult.FIRST_ORDERED_NODE_TYPE, null).singleNodeValue;

        function calculateStatistics() {
            // Get all td elements with aria-selected="true"
            const selectedCells = doc.querySelectorAll('td[aria-selected="true"]');
            
            // Convert inner text of each cell to a number and filter out NaN values
            const values = Array.from(selectedCells).map(cell => parseFloat(cell.innerText)).filter(num => !isNaN(num));

            if (values.length < 2) {
                container.style.display = 'none';
                return;
            }

            const sum = values.reduce((a, b) => a + b, 0);

            document.getElementById('maxValue').innerText = (Math.max(...values) || 0).toFixed(2);
            document.getElementById('minValue').innerText = (Math.min(...values) || 0).toFixed(2);
            document.getElementById('sumValue').innerText = (sum || 0).toFixed(2);
            document.getElementById('meanValue').innerText = (sum / values.length || 0).toFixed(2);
            
            container.style.display = 'block';
        }

        // Listen for mouseup event to trigger calculation
        doc.addEventListener('mouseup', () => {
            clearTimeout(calculationTimeout);
            calculationTimeout = setTimeout(calculateStatistics, 200);  // Adjust the delay as needed
        });

    </script>

    <style>
        .stats-header {
            background-color: rgba(255, 0, 0, 0.8); /* Red header background with transparency */
            color: white;               /* Header text color */
            padding: 10px;              /* Padding around header */
            font-weight: bold;          /* Bold header text */
            text-align: center;         /* Centered text */
            border-radius: 5px;        /* Rounded corners */
        }
        .stats-row {
            display: flex;              /* Use flexbox for horizontal layout */
            justify-content: space-between; /* Space items evenly */
            padding: 5px 0;            /* Padding for rows */
            border-bottom: 1px solid rgba(200, 200, 200, 0.5); /* Bottom border with transparency */
        }
        .stats-cell {
            flex: 1;                   /* Flex to fill space */
            text-align: center;        /* Centered text */
            padding: 5px;             /* Padding in cells */
            color: white;             /* White text for contrast */
        }
        .stats-row:last-child {
            border-bottom: none;       /* Remove bottom border from last row */
        }
        .stats-row:hover {
            background-color: rgba(255, 0, 0, 0.2); /* Change background on hover with transparency */
        }
    </style>
    """
    # Display the HTML component with JavaScript in Streamlit
    components.html(html_code, height=75, width=625)
