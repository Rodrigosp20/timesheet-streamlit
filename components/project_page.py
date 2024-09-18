import re
import streamlit as st
from utils import *
from io import BytesIO
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.properties import CalcProperties
from dateutil.relativedelta import relativedelta
from openpyxl.worksheet.formula import ArrayFormula
from datetime import timedelta


@st.cache_data
def get_salary_table(data):
    data = data.sort_values(by='date')
    data = data.query('Salário != 0')
    salary_changes = data[data['Salário'].diff() !=0 ]
    return list(zip(salary_changes['date'], salary_changes['Salário']))

def update_sheet_references(workbook, map:dict):
    # Pattern to match sheet references in formulas

    for sheet in ['Aux', 'Imputação Horas']:
        ws = workbook[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Check if cell contains a formula
                    formula = cell.value
                    
                    for old_name, new_name in map.items():
                        pattern = re.compile(rf'\b{old_name}\b')
                        # Get the formula from the cell's value
                        # Replace old sheet name with new sheet name in the formula
                        formula = pattern.sub(new_name, formula)
                        # Set the updated formula back to the cell

                    cell.value = formula

def generate_timesheet(project, start_date, end_date):
    wb = load_workbook("assets/Timesheet.xlsx")
         

    acts_df = st.session_state.activities.query('project == @project["name"]')
    
    filter_start = get_first_date(start_date)
    filter_end = get_first_date(end_date)
    
    wp_lines = [10, 141, 272, 403, 534, 665, 796, 927, 1058, 1189]
    act_line = 3

    ws = wb['Cronograma']
    ws['G9'] = project['start_date']

    assert len(acts_df['wp'].unique()) <= 10
    
    acts = {}

    ## Fill Cronograma
    for i, wp in enumerate(acts_df['wp'].unique()):
        
        wp_index = wp_lines[i]
        ws[f'B{wp_index}'] = wp
        acts[wp] = []

        for j, act in enumerate(acts_df.query('wp == @wp').itertuples(index=False)):
            
            act_index = wp_index + 1 + 13*j
            ws[f'B{act_index}'] = act.activity
            acts[wp].append(act.activity)

            ws[f'F{act_index+1}'].value = act.trl.split()[1]
            
            for col in range(7, 55):
                date = get_first_date(pd.Timestamp(project['start_date'] + relativedelta(months= col-7)))

                if  compare(date, act.start_date, act.end_date):
                    
                    if compare(date, act.real_start_date, act.real_end_date):
                        ws.cell(act_index+1, col).style =  'Ativo'
                    else:
                        ws.cell(act_index+1, col).style =  'Planeado'
                    
                else:

                    if compare(date, act.real_start_date, act.real_end_date):
                        ws.cell(act_index+1, col).style =  'Real'
        
        for i in range(act_index + 2, 1321):
            ws.row_dimensions[i].hidden = True

    map = {}
    
    ws = wb['Equipa de projeto']
    contracts = st.session_state.contracts.query('project == @project["name"]')
    
    #Fill Equipa de Projeto
    for line, contract in enumerate(contracts.itertuples(index=False), start=9):
        sheet_ind = line - 8
        
        ws[f'C{line}'] = contract.profile
        ws[f'D{line}'] = contract.person
        ws[f'E{line}'] = contract.gender
        ws[f'F{line}'] = contract.start_date

        #Get Profile data
        sh = st.session_state.sheets.query('person == @contract.person and date >= @get_first_date(@contract.start_date) and date <= @get_first_date(@contract.end_date)')
        planned = st.session_state.planned_work.query('person == @contract.person and project == @project["name"] and date >= @get_first_date(@contract.start_date) and date <= @get_first_date(@contract.end_date)')
        real = st.session_state.real_work.query('person == @contract.person and date >= @get_first_date(@contract.start_date) and date <= @get_first_date(@contract.end_date)').sort_values(by=['project', 'date'])
        working_days = st.session_state.working_days.query('project == @project["name"]')

        salary_list = get_salary_table(sh)
        for i,data in enumerate(salary_list):
            ws.cell(line, 13 + i*2).value = data[1]
            ws.cell(line, 14 + i*2).value = data[0]

        sheet = wb[f'{sheet_ind}. TBD']
        sheet.title = f'{sheet_ind}. {contract.person}'
        sheet.sheet_state = 'visible'

        sheet['D129'] = 'Horas Reais'

        map[f'{sheet_ind}. TBD'] = f'{sheet_ind}. {contract.person}'

        for col in sheet.iter_cols(min_col=1, max_col=55, min_row=1, max_row=1):
            for cell in col:
                cell.value = ""

        for i, work in enumerate(real.query('project != @project["name"]')['project'].unique(), start=124):
            if i==128:
                break

            sheet[f'D{i}'] = work

        # Fill Profile Sheet
        for i,col in enumerate(range(column_index_from_string('E'), column_index_from_string('BA'))):
            date = get_first_date(start_date + relativedelta(months=i))

            
            if compare(date, filter_start, filter_end):
                sheet.column_dimensions[get_column_letter(col)].hidden = False
            else:
                sheet.column_dimensions[get_column_letter(col)].hidden = True

            c_sh = sh.query('date == @date')
            
            if not c_sh.empty:

                sheet.cell(5, col).value = c_sh.iloc[0]['Jornada Diária']
                sheet.cell(8, col).value = c_sh.iloc[0]['Faltas']
                sheet.cell(9, col).value = c_sh.iloc[0]['Férias']

            wd_sh = working_days.query("date == @date")
            if not wd_sh.empty:
                sheet.cell(6, col).value = wd_sh.iloc[0]['day']
            else:
                sheet.cell(6, col).value = ''
            
            p_sh = planned.query('date == @date')
            if not p_sh.empty:

                for i, act_list in enumerate(acts.values()):
                    line = 15 + i*11

                    for line,act in enumerate(act_list,line):
                        if not (act_row := p_sh.query('activity == @act')).empty:
                            sheet.cell(line, col).value = act_row.iloc[0]['hours']
                
                for row in range(line+1,124):
                    sheet.row_dimensions[row].hidden= True
            
            r_sh = real.query('project != @project["name"] and date == @date')
            if not r_sh.empty:

                for line, work in enumerate(r_sh.itertuples(index=False), start=124):
                    sheet.cell(line,col).value = work.hours

            pr_sh = real.query('project == @project["name"] and date == @date') 
            if not pr_sh.empty:
                sheet.cell(129, col).value = pr_sh.iloc[0]["hours"]   
            
            for row in range(130,332):
                sheet.row_dimensions[row].hidden= True
        
    for i in range(sheet_ind + 1, 61):
        ws = wb[f'{i}. TBD']
        
        for col in ws.iter_cols(min_col=5, max_col=52, min_row=6, max_row=6):
            for cell in col:
                cell.value = ""
        
    update_sheet_references(wb, map)

    return wb

def find_person_sheet(name, sheets):
    for sheet in sheets:
        if name.lower() in sheet.lower():
            return sheet
    
    return None

def update_project(project, file, other_activities = True):

    contracts = st.session_state.contracts.query('project == @project["name"]')
    activities = st.session_state.activities.query('project == @project["name"]')
    
    sheets_name = pd.ExcelFile(file).sheet_names

    first_sheet = True
    sheets = pd.DataFrame()
    planned_works = pd.DataFrame()
    real_works = pd.DataFrame()
 
    for contract in contracts.itertuples(index=False):

        if not (sheet := find_person_sheet(contract.person, sheets_name)):
            continue

        df = pd.read_excel(file, sheet_name=sheet, header=3, usecols="D:AZ")

        df = df.rename(columns={df.columns[0]: 'date'})
        df = df.set_index("date")


        date_range = (df.columns >= pd.to_datetime(contract.start_date)) & (df.columns <= pd.to_datetime(contract.end_date))
        contract_range = df.columns[date_range]

        try:
            sheet = df.loc[['Jornada diária', 'N.º de dias \núteis','Faltas (horas/mês)','Férias (horas/mês)','Salário atualizado (€)','SS'], contract_range].fillna(0)
        except:
            raise Exception(f"Folha {sheet}: Erro na leitura da folha de horas")
        
        sheet = sheet.transpose().reset_index(names="date")
        
        sheet = sheet.rename(columns={
            "Jornada diária":"Jornada Diária",
            "N.º de dias \núteis":"day",
            "Faltas (horas/mês)":"Faltas",
            "Férias (horas/mês)":"Férias",
            "Salário atualizado (€)":"Salário"
        })

        try:
            sheet[['Jornada Diária', 'day', 'Faltas', 'Férias', 'Salário']] = sheet[['Jornada Diária', 'day', 'Faltas', 'Férias', 'Salário']].astype('float64')
        except:
            raise Exception(f"Erro a ler a folha {sheet} - Não foi possível converter a folha com Números Inválidos")
    
        sheet["SS"] = sheet["SS"] * 100
        sheet["person"] = contract.person
        sheet["date"] = pd.to_datetime(sheet['date'])

        if first_sheet:
            working_days = sheet[["date", "day"]]
            first_sheet = False
        
        sheet.drop('day', axis='columns', inplace=True)
        sheets = pd.concat([sheets, sheet])
        
        try:
            planned_work = df.loc[activities['activity'].unique(), contract_range]
        except:
            raise Exception(f"Folha {sheet}: Não foram encontradas as atividades")
        
        planned_work = planned_work[~planned_work.index.duplicated()].fillna(0)

        planned_work = planned_work.reset_index(names="activity").melt(id_vars='activity', var_name='date', value_name='hours')
        planned_work['person'] = contract.person
        planned_work['project'] = project["name"]
        
        try:
            planned_work["date"] = pd.to_datetime(planned_work['date']) 
        except:
            raise Exception(f"Folha {sheet}: Erro a converter datas da folha")

        planned_works = pd.concat([planned_works, planned_work])
        
 
        try:
            real_work = df.loc[df.index.str.contains('Horas Reais').fillna(False), contract_range]
        except:
            raise Exception(f"Folha {sheet}: Não foi encontrado as Horas Reais [Horas Reais PRR]")

        try:
            index_position = df.index.get_loc('Outras atividades')
            other_activities = df.iloc[index_position-3:index_position].loc[:, contract_range]
            other_activities= other_activities[other_activities.index.notna()]
        except:
            raise Exception(f"Folha {sheet}: Erro a ler outras atividades")

        real_work = pd.concat([real_work, other_activities]).fillna(0)
        real_work = real_work.reset_index(names="project").melt(id_vars='project', var_name='date', value_name='hours')

        real_work['person'] = contract.person
        real_work['date'] = pd.to_datetime(real_work['date'])
        real_work.loc[real_work['project'].str.contains('Horas Reais'), 'project'] = project["name"]

        real_works = pd.concat([real_works, real_work])
    
    st.session_state.sheets = pd.concat([st.session_state.sheets , sheets])
    st.session_state.sheets = st.session_state.sheets.drop_duplicates(subset=["person","date"], keep='last')

    st.session_state.real_work = pd.concat([st.session_state.real_work , real_works])
    st.session_state.real_work = st.session_state.real_work.drop_duplicates(subset=["person","project","date"], keep='last')

    st.session_state.working_days = pd.concat([st.session_state.working_days, working_days])
    st.session_state.working_days = st.session_state.working_days.drop_duplicates(subset=["project","date"], keep='last')

    for act in activities.itertuples(index=False):
        planned_works = planned_works.query('~ (activity == @act.activity and (date < @act.real_start_date or date > @act.real_end_date))')
    
    st.session_state.planned_work = pd.concat([st.session_state.planned_work , planned_works])
    st.session_state.planned_work = st.session_state.planned_work.drop_duplicates(subset=["project","person","activity","date"], keep='last')
       
    set_notification("success", "Projeto atualizado com sucesso", force_reset=True)

def delete_project(project):
    st.session_state.projects = st.session_state.projects.query('name != @project["name"]')
    st.session_state.activities = st.session_state.activities.query('project != @project["name"]')
    st.session_state.contracts = st.session_state.contracts.query('project != @project["name"]')
    st.session_state.working_days = st.session_state.contracts.query('project != @project["name"]')
    st.session_state.real_work = st.session_state.real_work.query('project != @project["name"]')
    st.session_state.planned_work = st.session_state.planned_work.query('project != @project["name"]')
    st.session_state.unsaved = True

def update_project_dates(project, start, end):
    st.session_state.projects.loc[st.session_state.projects['name'] == project["name"], ["start_date", "end_date"]] = [start, end]

    project_range = pd.date_range(start= get_first_date(start), end= end, freq='MS')
    business_days = []
    for month_start in project_range:
        business_days.append(len(pd.date_range(start=month_start, end=month_start + pd.offsets.MonthEnd(), freq=pd.offsets.BDay())))
    
    st.session_state.working_days = pd.concat([st.session_state.working_days, pd.DataFrame({"project": project["name"], "day":business_days, "date":project_range})])
    st.session_state.working_days = st.session_state.working_days.drop_duplicates(subset=["project", "date"])

def get_salary_info(person, start, end):

    sheet = st.session_state.sheets.query('person == @person and date >= @start and date <= @end').sort_values(by="date")
    
    previous_value = -1
    salary = []
    for d in sheet.itertuples(index=False):
        value = d.Salário

        if value != previous_value:
            salary.append((value, d.date))

        previous_value = value

    return salary

def generate_sheets(project, start, end):
    
    wb = load_workbook('assets/Sheet_Template.xlsx')
    template = wb['sheet']

    start = get_first_date(start)
    end = get_last_date(end)
    
    initial_diff = (start.year - project["start_date"].year) * 12 + start.month - project["start_date"].month
    
    for i in range(5, initial_diff + 5):
        template.column_dimensions[get_column_letter(i)].hidden = True
    
    x = initial_diff + 5
    for date in pd.date_range(start=start, end=end, freq="MS"):
        template.cell(row=4, column=x, value=date.strftime("%b/%y"))
        x += 1

    for i in range(x, 65):
        template.column_dimensions[get_column_letter(i)].hidden = True
        template.column_dimensions[get_column_letter(i)].width = 0

    for i in range(i, 200):
        template.column_dimensions[get_column_letter(i)].hidden

    contracts = st.session_state.contracts.query('project == @project["name"]')
    working_days = st.session_state.working_days.query('project == @project["name"]')
    
    person_line = 9
    for contract in contracts.itertuples(index=False):
        team_ws = wb['Equipa de projeto']

        salary = get_salary_info(contract.person, project["start_date"], project["end_date"])
        team_ws.cell(row=person_line, column=3, value= contract.profile)
        team_ws.cell(row=person_line, column=4, value= contract.person)
        team_ws.cell(row=person_line, column=5, value= contract.gender)
        team_ws.cell(row=person_line, column=6, value= contract.start_date)
        
        if project["end_date"] != contract.end_date:
            team_ws.cell(row=person_line, column=7, value= contract.end_date)

        for i, info in enumerate(salary, start=0):
            j = 2 * i

            team_ws.cell(row=person_line, column=9 + j, value= info[0])
            team_ws.cell(row=person_line, column=9 + j + 1, value= info[1])

        sheet = wb.copy_worksheet(template)

        sheet.title= contract.person

        df = st.session_state.sheets.query('person == @contract.person and date >= @start and date <= @end')
        df = df.merge(working_days[['date', 'day']], on="date", how="left").rename(columns={'day':'Dias Úteis'})
        real_work = st.session_state.real_work.query('person == @contract.person and date >= @start and date <= @end')

        project_work = real_work.query('project == @project["name"]')
        other_work = real_work.query('project != @project["name"]')

        other_work = other_work.sort_values(by="project")

        sheet["A1"] = contract.person

        row = 15
        for other_project in other_work['project'].unique():
            sheet.cell(row=row, column=4, value=other_project)
            row += 1
        
        for x in range(row, 21):
            sheet.row_dimensions[x].hidden = True

        planned_work = st.session_state.planned_work.query('person == @contract.person and project == @project["name"] and date >= @start and date <= @end')

        df_t = df.drop(columns='person').set_index('date')
        df_t = df_t.transpose()

        planned_work = planned_work.pivot(index="activity", columns="date", values="hours")
      
        horas_trabalhaveis = (df_t.loc['Jornada Diária'] * df_t.loc['Dias Úteis']).fillna(0)
        sum_wp = planned_work.sum()
        
        sum_wp= sum_wp.replace(0, np.nan)
        horas_trabalhaveis = horas_trabalhaveis.replace(0, np.nan)
        real_work = project_work.pivot(index="person", columns="date", values="hours")
    
        planned_work = ( (planned_work /sum_wp * real_work.loc[contract.person]).div(horas_trabalhaveis) ).fillna(0)
        planned_work = planned_work.reset_index(names="activity")

        planned_work = planned_work.merge(st.session_state.activities[['activity', "wp", 'trl']], on="activity", how="left")
        planned_work = planned_work.groupby(["wp", "trl"]).sum().drop(columns="activity")

        planned_work.columns = planned_work.columns.map(lambda col: col.strftime("%b/%y"))
        row = 23
        for wp in planned_work.index.get_level_values(0).unique():
            sheet.cell(row=row, column=3, value=wp)
            row += 2

        for row in range(row, 43):
            sheet.row_dimensions[row].hidden = True

        x = initial_diff + 5
        for date in pd.date_range(start=start, end=end, freq="MS"):
            if not (row := df.loc[df['date'] == date]).empty:
                sheet.cell(row=5, column=x, value= row['Jornada Diária'].iloc[0])
                sheet.cell(row=6, column=x, value= row['Dias Úteis'].iloc[0])
                sheet.cell(row=8, column=x, value= val if (val:= row['Faltas'].iloc[0]) != 0 else "")
                sheet.cell(row=9, column=x, value= val if (val:= row['Férias'].iloc[0]) != 0 else "")

            if not (val:= project_work.loc[project_work['date'] == date, 'hours']).empty:
                sheet.cell(row=14, column=x, value= val.iloc[0])

            row = 15
            for res in other_work.loc[other_work['date'] == date].itertuples(index=False):
                sheet.cell(row=row, column=x, value= res.hours)
                row += 1
            
            row = 23
            while (wp:= sheet.cell(row= row, column=3).value):
                try:
                    sheet.cell(row=row, column=x, value=planned_work.loc[(wp, 'TRL 3-4'), date.strftime("%b/%y")])
                except:
                    pass 
                try:
                    sheet.cell(row=row+1, column=x, value=planned_work.loc[(wp, 'TRL 5-9'), date.strftime("%b/%y")])
                except:
                    pass

                row += 2

            x += 1
            
        
        sheet.sheet_view.showGridLines = False

        blueFill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        blueFont = Font(color="0070C0", name="Aptos Narrow", bold=True)
        sheet.conditional_formatting.add('E23:BL42', CellIsRule('greaterThan', formula=['0'], fill=blueFill, font=blueFont))
    
        person_line +=1

    for row in range(person_line, 68 + 1):
        team_ws.row_dimensions[row].hidden = True

    del wb["sheet"]

    return wb

def generate_input(project, start_date, end_date):
    wb = load_workbook('assets/Timesheet - Input.xlsx')
    template = wb['sheet']

    filter_start = get_first_date(start_date)
    filter_end = get_first_date(end_date)

    start = get_first_date(project["start_date"])
    end = get_last_date(project["end_date"])
    
    template['E4'] = start

    project_contracts = st.session_state.contracts.query('project == @project["name"]')
    project_working_days = st.session_state.working_days.query('project == @project["name"]')
    project_sheets = st.session_state.sheets.query('person in @project_contracts["person"].unique() and date >= @start and date <= @end').merge(project_working_days[["date","day"]], on="date", how="left")
    project_planned_work = st.session_state.planned_work.query('project == @project["name"]').merge(st.session_state.activities[['activity', "wp", 'trl']], on="activity", how="left")
    project_real_work = st.session_state.real_work.query('project == @project["name"] and date >= @start and date <= @end')

    acts = []
    project_planned_work = project_planned_work.sort_values(by="activity")
    

    #Calculate Real Hours p/ Activity    
    sum_wp = project_planned_work.groupby(['person', 'wp', 'date'])['hours'].sum().reset_index().rename(columns={'hours':'wp_total_hours'})
    project_planned_work = pd.merge(project_planned_work, sum_wp, how="left", on=["person", "wp", "date"]).rename(columns={"hours":"planned_hours"})
    
    project_planned_work = pd.merge(project_planned_work, project_real_work[["person", "date", "hours"]], how="left", on=["person", "date"]).rename(columns={"hours":"real_hours"})
    project_planned_work["real_hours"] = (project_planned_work["planned_hours"] / project_planned_work["wp_total_hours"].replace(0, np.nan) * project_planned_work["real_hours"]).fillna(0)


    # Write Activities Names
    for offset, wp in enumerate(project_planned_work["wp"].unique()):
        line = 14 + 11 * offset

        template[f'D{line}'] = wp
        template[f'D{line+112}'] = wp
        wp_acts = []

        for i, act in enumerate(project_planned_work.query('wp == @wp').drop_duplicates(subset="activity").itertuples(index=False), start=1):
            template[f'D{line + i}'] = act.activity
            template[f'C{line + i}'] = act.trl

            template[f'D{line + i + 112}'] = act.activity
            template[f'C{line + i + 112}'] = act.trl
            wp_acts.append(act.activity)

        acts.append(wp_acts)
    
    # Hide Unused Acitivty Slots
    for i in range(line + 11, 124):

        template.row_dimensions[i].hidden = True
        template.row_dimensions[i + 112].hidden = True

    #Hide filtered Dates Column
    for i,col in enumerate(range(column_index_from_string('E'), column_index_from_string('BA'))):
        date = start_date + relativedelta(months=i)
        
        if compare(date, filter_start, filter_end):
            template.column_dimensions[get_column_letter(col)].hidden = False
        else:
            template.column_dimensions[get_column_letter(col)].hidden = True


    team_ws = wb['Equipa de projeto']

    for person_line, contract in enumerate(project_contracts.itertuples(index=False),9):

        df = project_sheets.query('person == @contract.person')
        df_planned = project_planned_work.query('person == @contract.person')

        salary = get_salary_table(df)

        team_ws.cell(row=person_line, column=3, value= contract.profile)
        team_ws.cell(row=person_line, column=4, value= contract.person)
        team_ws.cell(row=person_line, column=5, value= contract.gender)
        team_ws.cell(row=person_line, column=6, value= contract.start_date)
        
        if project["end_date"] != contract.end_date:
            team_ws.cell(row=person_line, column=7, value= contract.end_date)

        for i, info in enumerate(salary, start=0):
            team_ws.cell(row=person_line, column=13 + 2 * i, value= info[1])
            team_ws.cell(row=person_line, column=14 + 2 * i, value= info[0])

        sheet = wb.copy_worksheet(template)
        sheet.title= f"{person_line-8}. {contract.person}"

        person_real_work = st.session_state.real_work.query('person == @contract.person and project != @project["name"] and date >= @start and date <= @end').sort_values(by="project")
        other_projects = person_real_work["project"].unique()
        for row, project_name in enumerate(other_projects, start=236):
            sheet.cell(row=row, column=4, value=project_name)


        # Worker Sheet Filling
        for col, date in enumerate(pd.date_range(start=start, end=end, freq="MS"), start=5):
            
            # wp_sheet = ((wp_sheet / sum_wp * modifications.loc['Horas Reais']) / horas_trabalhaveis ).fillna(0)
            sheet.cell(row=1, column=1 , value= contract.person)

            if not (row := df.query('date == @date')).empty:

                sheet.cell(row=5, column=col, value= row['Jornada Diária'].iloc[0])
                sheet.cell(row=6, column=col, value= row['day'].iloc[0])
                sheet.cell(row=8, column=col, value= val if (val:= row['Faltas'].iloc[0]) != 0 else "")
                sheet.cell(row=9, column=col, value= val if (val:= row['Férias'].iloc[0]) != 0 else "")

            for offset, wp_acts in enumerate(acts):

                for i, act in enumerate(wp_acts):

                    line = 15 + offset * 11 + i

                    if not (row:= df_planned.query('activity == @act and date == @date')).empty:

                        sheet.cell(row=line, column=col, value=val if (val:= row['planned_hours'].iloc[0]) != 0 else "")
                        sheet.cell(row=line + 112, column=col, value=val if (val:= row['real_hours'].iloc[0]) != 0 else "")

            for line, project_name in enumerate(other_projects,start=236):
                if not (row:= person_real_work.query('project == @project_name and date == @date')).empty:
                    sheet.cell(row=line, column=col, value= val if (val:= row['hours'].iloc[0]) != 0 else "")
                
                

        sheet.sheet_view.showGridLines = False
    
    del wb["sheet"]

    return wb

def generate_pay_sheets(project, file, order_by, start, end, df_team, df_trl):
    start = get_first_date(start)
    end = get_last_date(end)

    sheets = st.session_state.sheets.query('person in @df_team["equipa"].unique() and date >= @start and date <= @end')
    working_days = st.session_state.working_days.query('project == @project["name"]')
    sheets = sheets.merge(working_days[['date', 'day']], on="date", how="left").rename(columns={'day':'Dias Úteis'})
    
    project_work = st.session_state.real_work.query('project == @project["name"] and date >= @start and date <= @end')
    project_work = project_work.rename(columns={'hours':'real_work'})
    planned_work = st.session_state.planned_work.query('project == @project["name"] and date >= @start and date <= @end')

    sheets['horas_trabalhaveis'] = sheets['Jornada Diária'] * sheets['Dias Úteis']
    
    planned_work = planned_work.merge(st.session_state.activities[['wp', 'trl', 'activity']], on="activity", how="left")
    
    sum_wp = planned_work.groupby(["person","date"])['hours'].sum().reset_index()
    sum_wp = sum_wp.rename(columns={'hours':'wp_sum'})    

    planned_work = planned_work.merge(project_work[["person", "date", "real_work"]], on=["person", "date"], how="left")
    planned_work = planned_work.merge(sum_wp[["person", "date", "wp_sum"]], on=["person", "date"], how="left")
    planned_work = planned_work.merge(sheets[["date", "person", "horas_trabalhaveis"]], on=["person", "date"], how="left")
    
    planned_work['res'] = ((planned_work['hours'] / planned_work['wp_sum'].replace(0, np.nan) * planned_work['real_work']).replace(0,np.nan) / planned_work['horas_trabalhaveis']).fillna(0)
    
    df_team = df_team.merge(planned_work[["person", "wp", "trl", "date", "res"]], left_on="equipa", right_on="person")
    df_team = df_team.groupby(["tecnico", "wp", "trl", "date"])["res"].sum().reset_index()
    df_team = df_team[df_team['res'] > 0]

    df_team = df_team.merge(df_trl, on=["wp", "trl"], how="left")
    df_team = df_team[~ pd.isna(df_team['code'])]
    
    if len(order_by) > 0:
        df_team = df_team.sort_values(by=order_by)

    wb = load_workbook(file)
    ws = wb['Mapa']

    for i, val in enumerate(df_team.itertuples(index=False), start=4):

        ws[f'C{i}'] = val.code
        ws[f'D{i}'] = val.tecnico
        ws[f'E{i}'] = '{}/{}'.format(val.date.month, val.date.year)
        ws[f'G{i}'] = val.res

    return wb

def project_widget(project):

    save, undo = get_topbar(project['name'])

    with st.container(border=True):
            
        start_date = st.date_input("Data de Inicio", key=f"project_date_initial_{st.session_state.key}", value=project['start_date'], format="DD/MM/YYYY", max_value=project['start_date'])
        end_date = st.date_input("Data de Termino", key=f"project_date_final_{st.session_state.key}", value=project['end_date'], format="DD/MM/YYYY", min_value=project['end_date'])

        if save:
            update_project_dates(project, start_date, end_date)

        if undo:
            reset_key()
            st.rerun()
                
        if st.button("Eliminar Projeto", key="delete_project",use_container_width=True):
            def action():
                delete_project(project)
                set_notification("success", "Projeto eliminado")

            get_dialog("Eliminar Projeto", "Se continuar o projeto será eliminado e não será possível recuperá-lo, continuar mesmo assim ?", action)

    with st.expander("Atualizar Projeto c/Timesheet"):
        sheet = st.file_uploader("Timesheet", type=".xlsx")
        
        if st.button("Atualizar", use_container_width=True, disabled=sheet is None):
            update_project(project, sheet)

    with st.expander("Gerar Folhas de Afetação"):
        start_date, end_date = st.slider(
            "Selecionar espaço temporal",
            min_value= project["start_date"],
            max_value= project["end_date"],
            value= (project["start_date"], project["end_date"]),
            format="MMMM/YYYY"
        )
        
        if st.button("Gerar Excel", use_container_width=True):
            
            wb = generate_sheets(project, start_date, end_date)

            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)
            
            save_excel(virtual_workbook.getvalue(), f"{project['name']}.xlsx")
    
    with st.expander("Gerar Folhas de Pagamentos"):
        start_date, end_date = st.slider(
            "Selecionar espaço temporal",
            min_value= project["start_date"],
            max_value= project["end_date"],
            value= (project["start_date"], project["end_date"]),
            format="MMMM/YYYY",
            key="slider_pay"
        )

        if template := st.file_uploader("Template", type=".xlsx", accept_multiple_files=False):
            df_team = pd.read_excel(template ,sheet_name="Referências", usecols="H", header=3, names=["tecnico"]).dropna()
            project_persons = st.session_state.contracts.query('project == @project["name"]')["person"]

            
            df_team = pd.concat([df_team, project_persons],axis=1)
            df_team.dropna(subset='tecnico', inplace=True)
            df_team.columns = ['tecnico','equipa']

            df_team = st.data_editor(
                df_team,
                column_config={
                    "equipa":st.column_config.SelectboxColumn(
                        "equipa",
                        options=st.session_state.contracts.query('project == @project["name"]')["person"].unique()
                    )
                },
                disabled=["tecnico"],
                use_container_width=True,
                hide_index=True
            )

            df_trl = pd.read_excel(template, sheet_name="Referências", usecols="E", header=3, names=["code"]).dropna()
            aux = st.session_state.inv_order_num.query('project == @project["name"]')
            df_trl = pd.merge(left=df_trl, right=aux[['code','wp','trl']], how='left', left_on='code', right_on='code')
            # df_trl['wp'] = None
            # df_trl['trl'] = None

            df_trl = st.data_editor(
                df_trl,
                column_config={
                    "code":st.column_config.TextColumn(
                        "Investimento"
                    ),
                    "wp":st.column_config.SelectboxColumn(
                        "wp",
                        options=st.session_state.activities.query('project == @project["name"]')['wp'].unique()
                    ),
                    "trl":st.column_config.SelectboxColumn(
                        "trl",
                        options=st.session_state.activities.query('project == @project["name"]')['trl'].unique()
                    )
                },
                disabled=['code'],
                hide_index=True,
                use_container_width=True
            )    

            options = {
                "Pessoa": "tecnico",
                "Nº Ordem": "code",
                "Data": "date"
            }

            order_by = st.multiselect('Ordernar Por', options.keys())         
            order_by = [options[option] for option in order_by]

        if st.button("Gerar Excel", key="pay_excel", use_container_width=True , disabled=False if template else True):
            df_trl['project'] = project["name"]
            
            st.session_state.inv_order_num.drop(st.session_state.inv_order_num.query('project == @project["name"]').index, inplace=True)
            st.session_state.inv_order_num = pd.concat([st.session_state.inv_order_num, df_trl]).reset_index(drop=True)

            st.session_state.unsaved = True

            wb = generate_pay_sheets(project, template, order_by, start_date, end_date, df_team, df_trl)
            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)

            save_excel(virtual_workbook.getvalue(), f"fpp_{project['name']}.xlsx")
    
    with st.expander("Exportar Timesheet"):

        client_version = st.toggle("Versão Cliente")
        
        filter_start, filter_end = st.slider("", min_value=get_first_date(project["start_date"]), max_value=get_first_date(project["end_date"]), format="MMMM/YYYY", value=(get_first_date(project["start_date"]), get_first_date(project["end_date"])), step=timedelta(weeks=4))
        
        if st.button("Gerar Excel", key="timesheet", use_container_width=True):
            
            if client_version:
                wb = generate_input(project, filter_start, filter_end)
            else:
                wb = generate_timesheet(project, filter_start, filter_end)

            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)

            save_excel(virtual_workbook.getvalue(), f"timesheet_{project['name']}{'_cliente' if client_version else ''}.xlsx")
    



