from io import BytesIO
import streamlit as st
from utils import *
from openpyxl.utils import get_column_letter


def generate_sheets(start_date, end_date):

    start_date = get_first_date(start_date)
    end_date = get_last_date(end_date)
    # df_filtered = st.session_state.real_work[~st.session_state.real_work['project'].str.contains("Horas previstas projeto", na=False)]
    sheets = st.session_state.sheets.query(
        'date >= @start_date and date <= @end_date').sort_values(by="date")
    days = st.session_state.working_days.query(
        'date >= @start_date and date <= @end_date').drop_duplicates(subset="date")

    wb = load_workbook('assets/Timesheet_Person_Template.xlsx')

    for person in sheets['person'].unique():
        ws = wb.copy_worksheet(wb['Person'])

        ws.title = person
        ws['B3'] = person

        work = st.session_state.real_work.query(
            'date >= @start_date and date <= @end_date and person == @person')

        # Write Project Names
        projects = work['project'].unique()
        for i, project in enumerate(projects, start=13):
            ws[f'B{i}'] = project

        for i in range(i, 22):
            ws.row_dimensions[i+1].hidden = True

        sheet = sheets.query('person == @person')

        for col_num, row in enumerate(sheet.itertuples(index=False), start=4):
            col = get_column_letter(col_num)

            ws[f'{col}5'] = row.date
            ws[f'{col}6'] = row._2
            ws[f'{col}7'] = days.loc[days["date"] == row.date, "day"].iloc[0]
            ws[f'{col}9'] = row.Faltas
            ws[f'{col}10'] = row.Férias

            date_work = work.query('date == @row.date')

            for line, project in enumerate(projects, start=13):
                try:
                    ws[f'{col}{line}'] = date_work.loc[date_work['project']
                                                       == project, "hours"].iloc[0]
                except:
                    pass

        for col in range(col_num, 51):
            ws.column_dimensions[get_column_letter(col+1)].hidden = True

        ws.sheet_view.showGridLines = False

    del wb["Person"]

    return wb


def update_workers(df: pd.DataFrame):

    if df['name'].nunique() != len(df):
        return set_notification("error", "Nome dos funcionários tem de ser único")

    for person in st.session_state.persons['name']:

        if person not in list(df['name']):

            st.session_state.sheets = st.session_state.sheets.query(
                'person != @person')
            st.session_state.real_work = st.session_state.real_work.query(
                'person != @person')
            st.session_state.planned_work = st.session_state.planned_work.query(
                'person != @person')
            st.session_state.contracts = st.session_state.contracts.query(
                'person != @person')
    # update workers
    st.session_state.persons = df

    # update genders in the contracts
    merged_df = pd.merge(st.session_state.contracts, df,
                         left_on='person', right_on='name', how='left', suffixes=('', '_person'))

    merged_df['gender'] = merged_df['gender_person']
    merged_df.drop(columns=['gender_person', 'name'], inplace=True)

    st.session_state.contracts = merged_df

    set_notification("success", "Funcionários atualizados com sucesso!")


def workers_widget():

    comp_name = st.session_state.company_name

    save, undo = get_topbar(f"Funcionários {comp_name}")

    tab1, tab2 = st.tabs(
        ["Funcionarios da Empresa", "Funcionários por Projeto"])

    st.session_state.persons = st.session_state.persons.reset_index(drop=True)
    df = tab1.data_editor(
        st.session_state.persons,
        column_config={
            "name": st.column_config.TextColumn("Nome", required=True),
            "gender": st.column_config.SelectboxColumn("Género", options=['M', 'F'], required=True, default="M")
        },
        num_rows='dynamic',
        hide_index=True,
        use_container_width=True,
        key=f"workers_list_{st.session_state.key}"
    )

    with tab1.expander("Gerar Folhas de Afetação p/ funcionario"):
        projects = st.session_state.projects
        min_date = projects["start_date"].min()
        max_date = projects["end_date"].max()

        if pd.notna(min_date) and pd.notna(max_date):
            start_date, end_date = st.slider(
                "Selecionar espaço temporal",
                min_value=min_date,
                max_value=max_date,
                value=(min_date, max_date),
                format="MMMM/YYYY",
                key="slider_func"
            )

            if st.button("Gerar Folha afetação"):
                wb = generate_sheets(start_date, end_date)

                virtual_workbook = BytesIO()
                wb.save(virtual_workbook)
                virtual_workbook.seek(0)

                save_excel(virtual_workbook.getvalue(), f"{comp_name}.xlsx")

    # Display the table in Streamlit
    pivot_df = st.session_state.contracts.pivot_table(
        index="project", columns="person", aggfunc="size", fill_value=0)
    pivot_df = pivot_df.applymap(lambda x: True if x > 0 else False)

    tab2.dataframe(pivot_df, use_container_width=True)

    if save:
        update_workers(df)

    if undo:
        reset_key()
